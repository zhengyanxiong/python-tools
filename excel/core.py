"""excel.core — Excel 解析/生成核心逻辑"""

from pathlib import Path


def parse_excel(file_path: str, sheet_name: str | None = None) -> list[list]:
    """解析 Excel 文件，返回二维列表。

    Args:
        file_path: Excel 文件路径
        sheet_name: 工作表名（默认第一个）

    Returns:
        [[row1_col1, row1_col2, ...], ...]
    """
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append(list(row))
    wb.close()
    return data


def excel_to_json(file_path: str, sheet_name: str | None = None) -> list[dict]:
    """解析 Excel 文件，第一行为表头，返回字典列表。"""
    rows = parse_excel(file_path, sheet_name)
    if not rows:
        return []
    headers = [str(h or f"col{i}") for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:]]


def generate_excel(file_path: str, headers: list[str], data: list[list]) -> str:
    """根据表头和数据生成 Excel 文件。"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for row in data:
        ws.append(row)
    wb.save(file_path)
    return str(Path(file_path).resolve())
